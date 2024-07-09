import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
import plotly.graph_objects as go


def get_token_price(token_name):
    """Retrieve token price from CoinMarketCap."""
    url = f'https://coinmarketcap.com/currencies/{token_name}'
    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        price_text = None

        for class_name in ['sc-d1ede7e3-0 hSTakI base-text', 'sc-d1ede7e3-0 fsQm base-text']:
            price_element = soup.find('span', {'class': class_name})
            if price_element:
                price_text = price_element.text.strip().replace('$', '').replace(',', '')
                break

        return float(price_text) if price_text else None
    except (requests.RequestException, ValueError) as e:
        print(f"Error fetching price for {token_name}: {e}")
        return None


def initialize_workbook(file_name, tokens):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(["DateTime", "TotalValue"] + [token for token in tokens] + [token + ' Last Quantity' for token in tokens])
        wb.save(file_name)
    return load_workbook(file_name)


def fetch_previous_quantities(ws, tokens):
    last_row = ws.max_row
    last_quantities = {}
    if last_row > 1:
        for col_index, token in enumerate(tokens):
            cell_value = ws.cell(row=last_row, column=3 + col_index + len(tokens)).value
            last_quantities[token] = float(cell_value) if cell_value is not None else 0
    else:
        last_quantities = {token: 0 for token in tokens}
    return last_quantities


def update_worksheet(wb, ws, tokens, token_values, current_quantities, file_name):
    total_value = sum(token_values.values())
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), total_value] + list(token_values.values()) + list(current_quantities.values()))
    wb.save(file_name)
    return total_value


def generate_hover_text(data, tokens):
    hover_texts = []
    for i, row in data.iterrows():
        hover_text = f"Date: {row['DateTime']}<br>Total Value: ${row['TotalValue']:.2f}"
        for token in tokens:
            token_last_quantity_col = token + ' Last Quantity'
            if token_last_quantity_col not in data.columns:
                continue

            if i == 0:
                quantity_change_text = f"{token} added with quantity {row[token_last_quantity_col]}"
            elif row[token_last_quantity_col] != data.iloc[i-1][token_last_quantity_col]:
                quantity_change_text = f"{token} quantity change: {data.iloc[i-1][token_last_quantity_col]} -> {row[token_last_quantity_col]}"
            else:
                quantity_change_text = ""

            if quantity_change_text:
                hover_text += f"<br>{quantity_change_text}"
                
        hover_texts.append(hover_text)
    return hover_texts


def create_plot(data, hover_texts):
    fig = go.Figure(data=[go.Scatter(x=data['DateTime'], y=data['TotalValue'], mode='lines+markers', name='Total Value', text=hover_texts, hoverinfo='text')])
    fig.update_layout(
        title='Total Value Over Time',
        xaxis_title='Date and Time',
        yaxis_title='Total Value ($)',
        dragmode='zoom',
        xaxis=dict(fixedrange=False),
        yaxis=dict(fixedrange=False)
    )
    fig.show()


def fetch_token_values(tokens, last_quantities):
    token_values = {}
    for token, quantity in tokens.items():
        price = get_token_price(token)
        token_values[token] = price * quantity if price else 0

        last_quantity = last_quantities[token]
        if last_quantity != quantity:
            print(f"Quantity change for {token}: was {last_quantity}, now {quantity}")
        elif last_quantity == 0:
            print(f"No previous quantity recorded for {token}. Added with quantity {quantity}")
    return token_values


def handle_prospective_prices(tokens):
    prospective_values = {}
    method = input("Do you want to enter a specific value or a multiplier for each token? (value/multiplier): ").strip().lower()
    for token, quantity in tokens.items():
        try:
            if method == "value":
                prospective_price = float(input(f"Enter prospective price for {token}: "))
            elif method == "multiplier":
                current_price = get_token_price(token) or 0
                multiplier = float(input(f"Enter multiplier for {token}: "))
                prospective_price = current_price * multiplier
            prospective_value = prospective_price * quantity
            prospective_values[token] = prospective_value
            print(f"Prospective value of {token} at {prospective_price}: ${prospective_value:.2f}")
        except ValueError:
            print(f"Invalid input for {token}. Skipping...")
    return prospective_values


def main():
	tokens = {
		"cardano": 20000,
		"bitcoin": 1,
		"solana": 200,
		"xrp": 6073.456088
	}

    file_name = 'ValueGraph.xlsx'
    wb = initialize_workbook(file_name, tokens)
    ws = wb.active

    last_quantities = fetch_previous_quantities(ws, tokens)
    current_quantities = tokens.copy()

    token_values = fetch_token_values(tokens, last_quantities)
    total_value = update_worksheet(wb, ws, tokens, token_values, current_quantities, file_name)

    if input("Do you want to enter prospective prices for your tokens? (yes/no): ").strip().lower() == "yes":
        handle_prospective_prices(tokens)

    data = pd.read_excel(file_name)
    hover_texts = generate_hover_text(data, tokens)
    create_plot(data, hover_texts)


if __name__ == "__main__":
    main()
